import re
import sys
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from run import main


@pytest.fixture
def sample_orders_csv(tmp_path: Path) -> Path:
    data = pd.DataFrame(
        [
            {
                "order_id": "A-001",
                "order_date": "2024-01-02",
                "status": "paid",
                "channel": "Email",
                "product": "Widget",
                "units": 2,
                "price": 100,
            },
            {
                "order_id": "A-002",
                "order_date": "2024-01-03",
                "status": "paid",
                "channel": "Social",
                "product": "Widget",
                "units": 1,
                "price": 50,
            },
            {
                "order_id": "A-003",
                "order_date": "2024-01-04",
                "status": "paid",
                "channel": "Email",
                "product": "Widget Pro",
                "units": 1,
                "price": 100,
            },
            {
                "order_id": "A-004",
                "order_date": "2024-01-09",
                "status": "paid",
                "channel": "Email",
                "product": "Widget",
                "units": 1,
                "price": 120,
            },
            {
                "order_id": "A-005",
                "order_date": "2024-01-10",
                "status": "paid",
                "channel": "Social",
                "product": "Widget",
                "units": 2,
                "price": 75,
            },
            {
                "order_id": "A-006",
                "order_date": "2024-01-11",
                "status": "paid",
                "channel": "Affiliate",
                "product": "Gadget",
                "units": 0,
                "price": 80,
            },
            {
                "order_id": "A-007",
                "order_date": "2024-01-16",
                "status": "paid",
                "channel": "Email",
                "product": "Widget",
                "units": 1,
                "price": 0,
            },
        ]
    )

    input_csv = tmp_path / "orders.csv"
    data.to_csv(input_csv, index=False)
    return input_csv


def test_main_generates_kpi_columns_and_expected_values(
    sample_orders_csv: Path,
    tmp_path: Path,
) -> None:
    outdir = tmp_path / "out"

    main(
        publish=False,
        input_path=sample_orders_csv,
        outdir=outdir,
        paid_only=True,
    )

    weekly_path = outdir / "weekly_summary.csv"
    weekly = pd.read_csv(weekly_path)
    clean_path = outdir / "clean_orders.csv"

    assert list(weekly.columns) == [
        "week",
        "channel",
        "orders",
        "units",
        "revenue",
        "aov",
        "revenue_wow_pct",
        "channel_revenue_share_pct",
    ]

    summary = weekly.set_index(["week", "channel"])

    # aov = revenue / orders (selected rows)
    assert summary.loc[("2024-01-02", "Email"), "aov"] == pytest.approx(150.0)
    assert summary.loc[("2024-01-09", "Social"), "aov"] == pytest.approx(150.0)

    # WoW growth uses previous week for same channel.
    # Email: week1=300, week2=120 => (120 - 300) / 300 * 100 = -60
    # week3=0, previous week for Email is 120 => (0 - 120) / 120 * 100 = -100
    assert summary.loc[("2024-01-09", "Email"), "revenue_wow_pct"] == pytest.approx(-60.0)
    assert summary.loc[("2024-01-16", "Email"), "revenue_wow_pct"] == pytest.approx(-100.0)

    # Revenue share should sum to ~100 for weeks with non-zero revenue.
    weekly_share = weekly.groupby("week")["channel_revenue_share_pct"].sum()
    weekly_revenue = weekly.groupby("week")["revenue"].sum()
    for week, share_total in weekly_share.items():
        if weekly_revenue.loc[week] > 0:
            assert share_total == pytest.approx(100.0, abs=0.01)
        else:
            assert share_total == pytest.approx(0.0, abs=0.01)

    # Output should be rounded and avoid long float artifacts in generated CSV text.
    raw_csv = weekly_path.read_text(encoding="utf-8")
    assert "806.4000000000001" not in raw_csv
    assert re.search(r"\d+\.\d{3,}", raw_csv) is None

    clean_csv = clean_path.read_text(encoding="utf-8")
    assert "300.00000000000006" not in clean_csv
    assert re.search(r"\d+\.\d{3,}", clean_csv) is None

    clean = pd.read_csv(clean_path)
    money_columns = [column for column in ["price", "_revenue"] if column in clean.columns]
    assert money_columns
    for column in money_columns:
        assert clean[column].map(lambda value: f"{value:.2f}").str.match(r"-?\d+\.\d{2}").all()

    # Excel Summary tab schema/value checks.
    workbook = load_workbook(outdir / "weekly_report.xlsx", data_only=True)
    sheet = workbook["Summary"]
    header = [cell.value for cell in sheet[1]]
    assert header == [
        "week",
        "channel",
        "orders",
        "units",
        "revenue",
        "aov",
        "revenue_wow_pct",
        "channel_revenue_share_pct",
    ]

    first_data_row = [cell.value for cell in sheet[2]]
    assert first_data_row[0] == "2024-01-02"
    assert first_data_row[1] == "Email"
    assert first_data_row[5] == pytest.approx(150.0)

    # Workbook usability traits should be configured on output sheets.
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref == f"A1:H{sheet.max_row}"

    column_widths = {
        column: sheet.column_dimensions[letter].width
        for letter, column in zip("ABCDEFGH", header)
    }
    assert all(width is not None for width in column_widths.values())
    assert all(10 <= width <= 45 for width in column_widths.values())

    # Long headers should have a non-default width driven by content sizing.
    assert column_widths["revenue_wow_pct"] > 14
    assert column_widths["channel_revenue_share_pct"] > 14
